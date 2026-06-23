import json
import logging
import threading
import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime
import os
import psycopg2
import psycopg2.extras
from logging.handlers import RotatingFileHandler
import requests
import traceback
import inspect

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_DISPONIVEL = True
except ImportError:
    EXCEL_DISPONIVEL = False

# ==========================================================
# CONFIGURAÇÕES
# ==========================================================

BASE_URL = "https://web.qualityautomacao.com.br/INTEGRACAO/"

SQL_TITULO_PAGAR = """
TRUNCATE TABLE titulo_pagar_post;

INSERT INTO titulo_pagar_post (
    fornecedorCodigo,tituloPagarCodigo, descricao,parcela,valor,dataMovimento,vencimento,
    planoContaGerencialCodigo,centroCustoCodigo,observacao,isDataValida,numerotitulo,
    empresaCodigo
)
SELECT
    fornecedorCodigo,tituloPagarCodigo,descricao,parcela,valor,dataMovimento,vencimento,
    planoContaGerencialCodigo,centroCustoCodigo,observacao,isDataValida,numerotitulo,
    empresaCodigo
FROM titulo_pagar_get;

update centro_custo_dest ccd
set centrocustocodigo_antigo = cco.centrocustocodigo
from centro_custo_ori cco
where trim(ccd.descricao)=trim(cco.descricao);

update fornecedor_dest fd
set fornecedorcodigo_antigo = fo.fornecedorcodigo
from fornecedor_ori fo
where trim(fd.cnpjcpf)=trim(fo.cnpjcpf);

update plano_conta_dest pcd
set planocontacodigo_antigo = pco.planocontacodigo
from plano_conta_ori pco
where trim(pcd.hierarquia)=trim(pco.hierarquia)
and trim(pcd.descricao)=trim(pco.descricao);

update titulo_pagar_post
set isdatavalida = true;

update titulo_pagar_post tpp
set fornecedorcodigo = fd.fornecedorcodigo
from fornecedor_dest fd
where tpp.fornecedorcodigo = fd.fornecedorcodigo_antigo;

update titulo_pagar_post tpp
set planocontagerencialcodigo = pcd.planocontacodigo
from plano_conta_dest pcd
where tpp.planocontagerencialcodigo = pcd.planocontacodigo_antigo;

update titulo_pagar_post tpp
set centrocustocodigo = ccd.centrocustocodigo
from centro_custo_dest ccd
where tpp.centrocustocodigo = ccd.centrocustocodigo_antigo;
"""
# Colunas esperadas por tabela — usadas para filtrar campos da API.
# Os nomes aqui são os "originais" retornados pela API (camelCase).
# O código busca o nome REAL da coluna no banco via information_schema,
# fazendo a correspondência case-insensitive automaticamente.

COLUNAS_ESPERADAS: dict[str, list[str]] = {
    "centro_custo_ori": [
        "centroCustoCodigo", "codigo", "descricao",
        "centroCustoCodigoExterno", "tipoCentroCusto",
    ],
    "centro_custo_dest": [
        "centroCustoCodigo", "codigo", "descricao",
        "centroCustoCodigoExterno", "tipoCentroCusto",
    ],
    "plano_conta_ori": [
        "planoContaCodigo", "codigo", "descricao",
        "hierarquia", "apuraDre", "natureza", "tipo",
    ],
    "plano_conta_dest": [
        "planoContaCodigo", "codigo", "descricao",
        "hierarquia", "apuraDre", "natureza", "tipo",
    ],
    "fornecedor_ori": [
        "fornecedorCodigo", "codigo", "fornecedorReferencia",
        "razao", "fantasia", "cnpjCpf", "contasFornecedor",
        "tipoPessoa", "fornecedorCodigoExterno",
    ],
    "fornecedor_dest": [
        "fornecedorCodigo", "codigo", "fornecedorReferencia",
        "razao", "fantasia", "cnpjCpf", "contasFornecedor",
        "tipoPessoa", "fornecedorCodigoExterno",
    ],
    "titulo_pagar_get": [
        "fornecedorCodigo", "tituloPagarCodigo", "descricao",
        "parcela", "dataMovimento", "vencimento",
        "planoContaGerencialCodigo", "centroCustoCodigo",
        "observacao", "isDataValida", "empresaCodigo",
        "notaEntradaCodigo", "dataPagamento", "situacao",
        "tipo", "tipoLancamento", "valor", "valorPago",
        "desconto", "acrescimo", "cheque", "dinheiro",
        "troco", "adiantamento", "cartao",
        "fornecedorIntermediadorCodigo", "numeroTitulo",
        "nomeFornecedor", "cpfCnpjFornecedor", "pagamento",
        "numeroRemessa", "planoContaGerencialNivel",
        "planoContaGerencialDescricao", "centroCustoDescricao",
        "quantidadeParcelas", "linhaDigitavel", "autorizado",
        "nossoNumero", "bancoFornecedor", "agenciaFornecedor",
        "contaFornecedor", "tipoChavePixFornecedor",
        "chavePixFornecedor", "qrCodePix", "tipoTributo",
        "codigoReceitaTributo", "renavam", "placa",
        "codigoMunicipio", "digitoFgts", "lacreConSocialFgts",
        "identificadorFgts", "codigo",
    ],
    "titulo_pagar_post": [
        "fornecedorCodigo", "tituloPagarCodigo", "descricao",
        "parcela", "dataMovimento", "vencimento",
        "planoContaGerencialCodigo", "centroCustoCodigo",
        "observacao", "isDataValida", "empresaCodigo",
    ]
}

# Colunas cujo tipo no banco é JSONB
COLUNAS_JSONB: set[str] = {"pagamento"}

# Mapeamento de campos: banco → API para Títulos a Pagar
MAPEAMENTO_TITULOS_POST = {
    "fornecedorCodigo": "fornecedorCodigo",
    "numeroTitulo": "numeroTitulo",
    "descricao": "descricao",
    "valor": "valorParcela",
    "parcela": "numeroParcelas",
    "dataMovimento": "dataMovimento",
    "vencimento": "vencimento",
    "planoContaGerencialCodigo": "planoContaCodigo",
    "centroCustoCodigo": "centroCustoCodigo",
    "observacao": "observacao",
    "isDataValida": "isDataValida",
}
ALIAS_COLUNAS: dict[str, dict[str, str]] = {
    "centro_custo_dest":  {"codigo": "centrocustocodigo_antigo"},
    "plano_conta_dest":   {"codigo": "planocontacodigo_antigo"},
    "fornecedor_dest":    {"codigo": "fornecedorcodigo_antigo"},
}
TABELAS_ORIGEM: dict[str, str] = {
    "CENTRO_CUSTO":          "centro_custo_ori",
    "PLANO_CONTA_GERENCIAL": "plano_conta_ori",
    "FORNECEDOR":            "fornecedor_ori",
    "TITULO_PAGAR":          "titulo_pagar_get",
}

TABELAS_DESTINO: dict[str, str] = {
    "CENTRO_CUSTO":          "centro_custo_dest",
    "PLANO_CONTA_GERENCIAL": "plano_conta_dest",
    "FORNECEDOR":            "fornecedor_dest",
    # TITULO_PAGAR não é executado no modo DESTINO
}

# ==========================================================
# LOG
# ==========================================================

PASTA_LOG_ERROS = r"C:\Quality\LOG\erros_implantacao"

os.makedirs(PASTA_LOG_ERROS, exist_ok=True)

arquivo_log = os.path.join(
    PASTA_LOG_ERROS,
    f"implantacao_{datetime.now().strftime('%Y%m%d')}.log"
)

log = logging.getLogger(__name__)
log.setLevel(logging.INFO)

formatter = logging.Formatter(
    "%(asctime)s - %(levelname)s - %(message)s"
)

# Console
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)

# Arquivo
file_handler = RotatingFileHandler(
    arquivo_log,
    maxBytes=5 * 1024 * 1024,  # 5 MB
    backupCount=10,
    encoding="utf-8"
)
file_handler.setFormatter(formatter)

log.addHandler(console_handler)
log.addHandler(file_handler)

# ==========================================================
# MAPEAMENTO DE COLUNAS (API camelCase → banco real)
# ==========================================================


class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Integração API → PostgreSQL")
        self.root.resizable(True, True)
        self.root.minsize(520, 420)
        self._centralizar(520, 420)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        self.conn_params: dict | None = None
        self.token_destino: str = ""          # token da filial de destino (fase 2)
        self.tela_conexao = TelaConexao(self.root, self)
        self.tela_principal = TelaPrincipal(self.root, self)
        self.tela_titulos_post: TelaTitulosPost | None = None

        self.show_conexao()

    def _centralizar(self, largura, altura):
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() - largura) // 2
        y = (self.root.winfo_screenheight() - altura) // 2
        self.root.geometry(f"{largura}x{altura}+{x}+{y}")

    def show_conexao(self):
        self.tela_principal.hide()
        self.tela_conexao.show()

    def show_principal(self, conn_params: dict):
        self.conn_params = conn_params
        self.tela_principal.set_conn_params(conn_params)
        self.tela_conexao.hide()
        self.tela_principal.show()

    def show_titulos_post(self, conn_params: dict, token_destino: str = ""):
        """Exibe a tela de títulos POST para envio à API."""
        if token_destino:
            self.token_destino = token_destino
        if self.tela_titulos_post is None:
            self.tela_titulos_post = TelaTitulosPost(self.root, self)
        self.tela_principal.hide()
        self.tela_titulos_post.carregar_titulos(conn_params)
        self.tela_titulos_post.show()

    def run(self):
        self.root.mainloop()


# ==========================================================
# MAPEAMENTO DE COLUNAS (API camelCase → banco real)
# ==========================================================

def validar_estrutura_banco(conn) -> list[str]:
    """
    Verifica se todas as tabelas e colunas mapeadas (incluindo aliases)
    estão presentes no banco. Retorna uma lista de strings com os erros.
    """
    erros = []
    tabelas_esperadas = list(COLUNAS_ESPERADAS.keys())
    
    with conn.cursor() as cur:
        for tabela in tabelas_esperadas:
            cur.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = %s
                )
            """, (tabela,))
            if not cur.fetchone()[0]:
                erros.append(f"Tabela ausente: {tabela}")
                continue
                
            colunas_reais = obter_colunas_reais(conn, tabela)
            colunas_api = COLUNAS_ESPERADAS[tabela]
            aliases = ALIAS_COLUNAS.get(tabela, {})
            
            for col_api in colunas_api:
                col_banco_esperada = aliases.get(col_api, col_api)
                if col_banco_esperada.lower() not in colunas_reais:
                    erros.append(f"{col_banco_esperada} (em {tabela})")
                    
    return erros


def obter_colunas_reais(conn, tabela: str) -> dict[str, str]:
    """
    Consulta o information_schema e retorna um dict:
        { nome_lower: nome_real_no_banco }
    onde nome_lower é o nome em minúsculas para comparação case-insensitive.
    """
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = %s
            ORDER BY ordinal_position
            """,
            (tabela,),
        )
        rows = cur.fetchall()
    return {r[0].lower(): r[0] for r in rows}


def mapear_colunas(conn, tabela: str, colunas_api: list[str],
                   log_fn=None) -> list[tuple[str, str]]:
    _log = log_fn or log.info
    reais = obter_colunas_reais(conn, tabela)
    aliases = ALIAS_COLUNAS.get(tabela, {})  # ← alias por tabela

    mapeamento = []
    for col_api in colunas_api:
        col_lookup = aliases.get(col_api, col_api)  # aplica alias se existir
        col_banco = reais.get(col_lookup.lower())
        if col_banco:
            mapeamento.append((col_api, col_banco))
        else:
            _log(f"  ⚠ Coluna '{col_api}' não encontrada em '{tabela}' — ignorada.")

    return mapeamento


# ==========================================================
# TELA DE CONEXÃO COM O BANCO
# ==========================================================

class TelaConexao:
    """Tela de configuração e teste de conexão ao PostgreSQL."""

    def __init__(self, root, app):
        self.root = root
        self.app = app
        self.frame = tk.Frame(root)
        self.frame.columnconfigure(1, weight=1)
        self._build()

    def _build(self):
        PAD = {"padx": 14, "pady": 6}

        tk.Label(self.frame, text="Tipo de Conexão:",
                 anchor="w", font=("Segoe UI", 10, "bold")).grid(
                     row=0, column=0, columnspan=2, sticky="w", **PAD)

        self.tipo_var = tk.StringVar(value="local")
        frame_tipo = tk.Frame(self.frame)
        frame_tipo.grid(row=1, column=0, columnspan=2, padx=18, sticky="w")
        tk.Radiobutton(frame_tipo, text="Conexão Local",
                       variable=self.tipo_var, value="local",
                       command=self._aplicar_defaults_local).pack(side="left", padx=6)
        tk.Radiobutton(frame_tipo, text="Conexão Remota",
                       variable=self.tipo_var, value="remota",
                       command=self._limpar_host).pack(side="left", padx=6)

        campos = [
            ("Hostname / IP:",  "entry_host",    False),
            ("Porta:",          "entry_porta",   False),
            ("Usuário:",        "entry_usuario", False),
            ("Senha:",          "entry_senha",   True),
            ("Nome do Banco:",  "entry_banco",   False),
        ]
        for i, (label, attr, oculto) in enumerate(campos, start=2):
            tk.Label(self.frame, text=label, anchor="w").grid(
                row=i, column=0, sticky="w", **PAD)
            if attr == "entry_banco":
                entry = ttk.Combobox(self.frame, width=32, values=[], state="normal")
            else:
                entry = tk.Entry(self.frame, width=34)
                if oculto:
                    entry.config(show="*")
            entry.grid(row=i, column=1, sticky="ew", padx=(0, 18), pady=4)
            setattr(self, attr, entry)

        self.entry_porta.bind("<FocusOut>", self._on_porta_change)
        self.entry_porta.bind("<Return>", self._on_porta_change)

        self.lbl_status = tk.Label(
            self.frame, text="", fg="gray", anchor="w",
            wraplength=460, justify="left",
        )
        self.lbl_status.grid(row=8, column=0, columnspan=2,
                              padx=18, pady=(6, 0), sticky="w")

        frame_btn = tk.Frame(self.frame)
        frame_btn.grid(row=9, column=0, columnspan=2, pady=16)

        btn_testar = tk.Button(frame_btn, text="Testar Conexão",
                  command=self._testar, width=16,
                  bg="#5C8A5C", fg="white", relief="flat")
        btn_conectar = tk.Button(frame_btn, text="Conectar",
                  command=self._conectar, width=16,
                  bg="#0078D7", fg="white", relief="flat")
        btn_cancelar = tk.Button(frame_btn, text="Sair",
                  command=self.root.quit, width=16,
                  relief="flat")

        btn_testar.pack(side="left", padx=6)
        btn_conectar.pack(side="left", padx=6)
        btn_cancelar.pack(side="left", padx=6)

        self._aplicar_defaults_local()

    def show(self):
        self.frame.grid(row=0, column=0, sticky="nsew")

    def hide(self):
        self.frame.grid_remove()

    def _aplicar_defaults_local(self):
        self.entry_host.delete(0, tk.END)
        self.entry_host.insert(0, "localhost")
        self.entry_porta.delete(0, tk.END)
        self.entry_porta.insert(0, "5432")
        self.entry_usuario.delete(0, tk.END)
        self.entry_usuario.insert(0, "postgres")

    def _limpar_host(self):
        self.entry_host.delete(0, tk.END)
        self.entry_porta.delete(0, tk.END)
        self.entry_porta.insert(0, "5432")

    def _coletar_params(self) -> dict | None:
        host    = self.entry_host.get().strip()
        porta   = self.entry_porta.get().strip()
        banco   = self.entry_banco.get().strip()
        usuario = self.entry_usuario.get().strip()
        senha   = self.entry_senha.get()

        erros = []
        if not host:
            erros.append("• Hostname / IP é obrigatório.")
        if not porta.isdigit():
            erros.append("• Porta deve ser numérica (default: 5432).")
        if not usuario:
            erros.append("• Usuário é obrigatório.")
        if not banco:
            erros.append("• Nome do Banco é obrigatório.")
        

        if erros:
            self._status("\n".join(erros), "red")
            return None

        return dict(host=host, port=int(porta),
                    dbname=banco, user=usuario, password=senha)

    def _on_porta_change(self, event=None):
        if not self.entry_host.get().strip() or not self.entry_porta.get().strip() or not self.entry_usuario.get().strip():
            return
        self._listar_bancos_disponiveis()

    def _listar_bancos_disponiveis(self):
        host = self.entry_host.get().strip()
        porta = self.entry_porta.get().strip()
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get()

        if not porta.isdigit():
            return

        self._status("Consultando bancos disponíveis…", "gray")
        try:
            conn = psycopg2.connect(
                host=host,
                port=int(porta),
                dbname="postgres",
                user=usuario,
                password=senha,
                connect_timeout=8,
            )
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT datname FROM pg_database WHERE datistemplate = false ORDER BY datname"
                )
                bancos = [row[0] for row in cur.fetchall()]
            conn.close()

            self.entry_banco["values"] = bancos
            if bancos:
                self.entry_banco.set(bancos[0])
                self._status(f"✔ Bancos carregados: {len(bancos)} disponíveis.", "green")
            else:
                self._status("✔ Conexão OK, mas nenhum banco disponível encontrado.", "orange")
        except psycopg2.OperationalError as exc:
            self._status(f"✘ Falha ao listar bancos: {_mensagem_amigavel(exc)}", "red")
        except Exception as exc:
            self._status(f"✘ Erro ao listar bancos: {exc}", "red")

    def _status(self, msg, cor="gray"):
        self.lbl_status.config(text=msg, fg=cor)
        self.root.update_idletasks()

    def _testar(self):
        host = self.entry_host.get().strip()
        porta = self.entry_porta.get().strip()
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get()

        erros = []
        if not host:
            erros.append("• Hostname / IP é obrigatório.")
        if not porta.isdigit():
            erros.append("• Porta deve ser numérica (default: 5432).")
        if not usuario:
            erros.append("• Usuário é obrigatório.")

        if erros:
            self._status("\n".join(erros), "red")
            return

        self._status("Testando conexão…", "gray")
        try:
            conn = psycopg2.connect(
                host=host,
                port=int(porta),
                dbname="postgres",
                user=usuario,
                password=senha,
                connect_timeout=8,
            )
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT datname FROM pg_database WHERE datistemplate = false ORDER BY datname"
                )
                bancos = [row[0] for row in cur.fetchall()]
            conn.close()

            self.entry_banco["values"] = bancos
            if bancos:
                self.entry_banco.set(bancos[0])
                self._status(f"✔ Conexão bem-sucedida! {len(bancos)} bancos encontrados.", "green")
            else:
                self._status("✔ Conexão bem-sucedida, mas nenhum banco encontrado.", "orange")
        except psycopg2.OperationalError as exc:
            self._status(f"✘ Falha: {_mensagem_amigavel(exc)}", "red")
        except Exception as exc:
            self._status(f"✘ Erro inesperado: {exc}", "red")

    def _conectar(self):
        params = self._coletar_params()
        if not params:
            return
        self._status("Conectando…", "gray")
        try:
            conn = psycopg2.connect(**params, connect_timeout=8)
            conn.close()
            self.app.show_principal(params)
        except psycopg2.OperationalError as exc:
            self._status(f"✘ Falha: {_mensagem_amigavel(exc)}", "red")
        except Exception as exc:
            self._status(f"✘ Erro inesperado: {exc}", "red")


# ==========================================================
# TELA PRINCIPAL
# ==========================================================

class TelaPrincipal:
    """Tela principal: token, flag, execução e progresso — fluxo em 2 fases."""

    _FASES = [
        {
            "fase":        1,
            "flag":        "GET_ORIGEM",
            "mapa":        "TABELAS_ORIGEM",   # resolvido em _iniciar_thread
            "titulo":      "Importação — Filial de Origem  (1 / 2)",
            "badge":       "▶  Fase 1 de 2  —  GET_ORIGEM",
            "label_token": "Token API — Filial de Origem:",
            "btn_text":    "Importar Origem",
            "msg_inicio":  "🔄  Iniciando importação da filial de ORIGEM…",
            "msg_ok":      "✅  Filial de ORIGEM importada com sucesso!",
        },
        {
            "fase":        2,
            "flag":        "GET_DESTINO",
            "mapa":        "TABELAS_DESTINO",
            "titulo":      "Importação — Filial de Destino  (2 / 2)",
            "badge":       "▶  Fase 2 de 2  —  GET_DESTINO",
            "label_token": "Token API — Filial de Destino:",
            "btn_text":    "Importar Destino",
            "msg_inicio":  "🔄  Iniciando importação da filial de DESTINO…",
            "msg_ok":      "✅  Filial de DESTINO importada com sucesso!",
        },
    ]

    def __init__(self, root, app):
        self.root = root
        self.app = app
        self.conn_params: dict | None = None
        self.frame = tk.Frame(root)
        self.frame.columnconfigure(0, weight=1)
        self.frame.columnconfigure(1, weight=1)
        self._fase_idx = 0
        self._build()

    # ------------------------------------------------------------------ #
    #  Build                                                               #
    # ------------------------------------------------------------------ #
    def _build(self):
        PAD = {"padx": 10, "pady": 6}

        # Título dinâmico
        self.lbl_titulo = tk.Label(
            self.frame, text="",
            anchor="center", font=("Segoe UI", 11, "bold"),
        )
        self.lbl_titulo.grid(row=0, column=0, columnspan=2, sticky="ew", **PAD)

        # Badge de fase
        self.lbl_badge = tk.Label(
            self.frame, text="", fg="#2E8B57",
            font=("Segoe UI", 9), anchor="center",
        )
        self.lbl_badge.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 4))

        # Label do token (muda por fase)
        self.lbl_token = tk.Label(
            self.frame, text="",
            anchor="center", font=("Segoe UI", 10, "bold"),
        )
        self.lbl_token.grid(row=2, column=0, columnspan=2, sticky="ew", **PAD)

        # Entrada do token
        self.entry_chave = tk.Entry(self.frame, width=32, justify="center")
        self.entry_chave.grid(row=3, column=0, columnspan=2,
                               padx=60, pady=4, sticky="ew")

        # Barra de progresso
        self.progress = ttk.Progressbar(
            self.frame, orient="horizontal", length=320, mode="determinate",
        )
        self.progress.grid(row=4, column=0, columnspan=2,
                           padx=60, pady=10, sticky="ew")

        # Área de log
        frame_log = tk.Frame(self.frame)
        frame_log.grid(row=5, column=0, columnspan=2, padx=18, sticky="nsew")
        self.frame.rowconfigure(5, weight=1)
        self.txt_log = tk.Text(
            frame_log, height=8, state="disabled",
            bg="#f4f4f4", relief="flat", font=("Consolas", 9),
        )
        self.txt_log.pack(side="left", fill="both", expand=True)
        scroll = tk.Scrollbar(frame_log, command=self.txt_log.yview)
        scroll.pack(side="right", fill="y")
        self.txt_log.config(yscrollcommand=scroll.set)

        # Botões
        frame_btn = tk.Frame(self.frame)
        frame_btn.grid(row=6, column=0, columnspan=2, pady=14)

        tk.Button(
            frame_btn, text="Voltar", command=self._voltar,
            width=12, relief="flat",
        ).pack(side="left", padx=6)

        # NOVO BOTÃO
        tk.Button(
            frame_btn,
            text="↺ Recarregar",
            command=self._recarregar,
            width=12,
            bg="#B8860B",
            fg="white",
            relief="flat",
        ).pack(side="left", padx=6)

        self.btn_executar = tk.Button(
            frame_btn, text="",
            command=self._iniciar_thread, width=14,
            bg="#2E8B57", fg="white", relief="flat",
        )
        self.btn_executar.pack(side="left", padx=6)

        # Aplica textos da fase inicial
        self._aplicar_fase()

    # ------------------------------------------------------------------ #
    #  RECARREGAR TELA 2                                                    #
    # ------------------------------------------------------------------ #
    def _recarregar(self):
        """Reinicia o processo desde a fase 1."""
        if not messagebox.askyesno(
            "Reiniciar Processo",
            "Deseja reiniciar todo o processo?"
        ):
            return

        self._fase_idx = 0

        self.entry_chave.delete(0, "end")

        self.progress["value"] = 0

        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.config(state="disabled")

        self.btn_executar.config(state="normal")

        self._aplicar_fase()

    # ------------------------------------------------------------------ #
    #  Controle de fase                                                    #
    # ------------------------------------------------------------------ #
    def _aplicar_fase(self):
        """Atualiza todos os textos e botões conforme _fase_idx."""
        cfg = self._FASES[self._fase_idx]
        self.lbl_titulo.config(text=cfg["titulo"])
        self.lbl_badge.config(text=cfg["badge"])
        self.lbl_token.config(text=cfg["label_token"])
        self.btn_executar.config(text=cfg["btn_text"])
        self.entry_chave.focus_set()

    def _reset_form(self):
        """Limpa entry, progresso e log — chamado entre as fases."""
        self.entry_chave.delete(0, "end")
        self.progress.config(value=0)
        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.config(state="disabled")

    def _concluir_fase(self):
        """Decide se avança de fase ou habilita navegação (fase 2 concluída)."""
        fase_concluida = self._fase_idx + 1

        if fase_concluida < len(self._FASES):
            # ── Fase 1 concluída → prepara fase 2 ──────────────────────
            messagebox.showinfo(
                "Fase 1 concluída",
                "✅ Filial de ORIGEM importada com sucesso!\n\n"
                "Clique OK para importar a Filial de DESTINO.",
            )
            self._fase_idx += 1
            self._reset_form()
            self._aplicar_fase()
            self.btn_executar.config(state="normal")
        else:

            self._log(
                "🔄 Executando preparação automática dos títulos..."
            )

            if not self._executar_sql_titulos():
                self.btn_executar.config(state="normal")
                return

            self._log(
                "✔ Preparação concluída."
            )

            self.root.after(
                0,
                lambda: messagebox.showinfo(
                    "Importação concluída",
                    "✅ Informações adicionadas no banco de dados com sucesso.\n\n"
                    "Pressione OK para visualizar os títulos."
                )
            )

            self.root.after(
                0,
                lambda: self.app.show_titulos_post(
                    self.app.conn_params,
                    token_destino=self.app.token_destino,
                )
            )

    # ------------------------------------------------------------------ #
    #  Show / Hide / Params                                                #
    # ------------------------------------------------------------------ #
    def show(self):
        self.frame.grid(row=0, column=0, sticky="nsew")

    def hide(self):
        self.frame.grid_remove()

    def set_conn_params(self, conn_params: dict):
        self.conn_params = conn_params

    # ------------------------------------------------------------------ #
    #  Log helper                                                          #
    # ------------------------------------------------------------------ #
    def _log(self, msg: str):
        log.info(msg)
        self.txt_log.config(state="normal")
        self.txt_log.insert("end", msg + "\n")
        self.txt_log.see("end")
        self.txt_log.config(state="disabled")
        self.root.update_idletasks()

    # ------------------------------------------------------------------ #
    #  Execução                                                            #
    # ------------------------------------------------------------------ #
    def _iniciar_thread(self):
        chave = self.entry_chave.get().strip()
        if not chave:
            messagebox.showwarning("Campo obrigatório",
                                   "Informe o Token antes de continuar.")
            return

        self.btn_executar.config(state="disabled")
        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.config(state="disabled")
        self.progress["value"] = 0

        cfg  = self._FASES[self._fase_idx]
        mapa = TABELAS_ORIGEM if cfg["flag"] == "GET_ORIGEM" else TABELAS_DESTINO

        # Salva o token do destino (fase 2) para reutilização na tela de Títulos
        if cfg["flag"] == "GET_DESTINO":
            self.app.token_destino = chave

        self._log(cfg["msg_inicio"])

        threading.Thread(
            target=self._executar,
            args=(chave, mapa),
            daemon=True,
        ).start()

    def _voltar(self):
        if self._fase_idx > 0:
            # Volta da Fase 2 para a Fase 1
            self._fase_idx -= 1
            self._reset_form()
            self._aplicar_fase()
            self.btn_executar.config(state="normal")
            # Limpa token de destino em memória
            if hasattr(self.app, "token_destino"):
                self.app.token_destino = ""
        else:
            # Volta da Fase 1 para a Conexão
            self._reset_form()
            self.app.show_conexao()

    def _mostrar_titulos(self):
        self.hide()
        self.app.show_titulos_post(
            self.app.conn_params
        )

    def _executar(self, chave: str, mapa: dict[str, str]):
        cfg    = self._FASES[self._fase_idx]
        total  = len(mapa)
        erros  = []
        erros_detalhados = {}

        try:
            conn = psycopg2.connect(**self.conn_params)
            self._log("✔ Conexão ao banco estabelecida.")

            # Validação da estrutura do banco (Apenas Fase 1)
            if cfg["flag"] == "GET_ORIGEM":
                self._log("🔍 Validando estrutura do banco de dados...")
                erros_bd = validar_estrutura_banco(conn)
                if erros_bd:
                    self._log("✘ Estrutura do banco incompatível ou incompleta.")
                    msg = "Foram identificadas colunas ou tabelas obrigatórias inexistentes na base selecionada:\n\n"
                    for err in erros_bd[:10]: # Mostra os primeiros 10 para não estourar a tela
                        msg += f"• {err}\n"
                    if len(erros_bd) > 10:
                        msg += f"... e mais {len(erros_bd) - 10} item(ns).\n"
                    
                    msg += "\nVerifique se o banco informado é o correto e retorne à tela de conexão para realizar uma nova conexão antes de prosseguir com a importação."
                    
                    self.root.after(0, lambda: messagebox.showerror(
                        "Erro ao validar a estrutura do banco de dados.",
                        msg
                    ))
                    self.root.after(0, lambda: self.btn_executar.config(state="normal"))
                    conn.close()
                    return
                self._log("✔ Estrutura validada com sucesso.")

            # Limpa tabelas apenas na Fase 1
            if cfg["flag"] == "GET_ORIGEM":

                tabelas_limpar = [
                    "centro_custo_ori",
                    "plano_conta_ori",
                    "fornecedor_ori",
                    "titulo_pagar_get",
                    "centro_custo_dest",
                    "plano_conta_dest",
                    "fornecedor_dest",
                    "titulo_pagar_post",
                ]

                with conn.cursor() as cur:
                    for tabela in tabelas_limpar:
                        self._log(f"🗑 Limpando tabela '{tabela}'...")
                        cur.execute(
                            f'TRUNCATE TABLE "{tabela}" RESTART IDENTITY CASCADE'
                        )

                conn.commit()
                self._log("✔ Todas as tabelas foram limpas.")
        except psycopg2.OperationalError as exc:

            caminho_log = registrar_erro(
                "Falha na conexão com PostgreSQL",
                exc
            )
            self._log(
                f"✘ Falha na conexão: {_mensagem_amigavel(exc)}"
            )
            
            msg = f"Falha na conexão:\n{_mensagem_amigavel(exc)}"
            if caminho_log:
                msg += f"\n\nArquivo de erro gerado:\n{caminho_log}\n\n"
                
            self.root.after(
                0,
                lambda m=msg: messagebox.showerror("Erro de Conexão", m)
            )
            
            self.root.after(
                0,
                lambda: self.btn_executar.config(state="normal")
            )
            return

        for i, (endpoint, tabela) in enumerate(mapa.items(), start=1):
            self._log(f"[{i}/{total}] Consultando endpoint {endpoint}…")
            try:
                registros = buscar_dados(endpoint, chave, log_fn=self._log)
                self._log(f"  → {len(registros)} registros recebidos da API.")

                if registros:
                    inseridos = inserir_no_banco(conn, tabela, registros,
                                                 log_fn=self._log)
                    self._log(f"  ✔ {inseridos} registros inseridos em '{tabela}'.")
                else:
                    self._log(f"  ⚠ Nenhum dado retornado para '{tabela}'.")

            except psycopg2.errors.UndefinedTable:
                msg = f"Tabela '{tabela}' não existe no banco."

                self._log(f"  ✘ {msg}")
                erros.append(f"{endpoint}: {msg}")
                erros_detalhados[endpoint] = ("UndefinedTable", None)

            except psycopg2.Error as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                msg = str(exc).splitlines()[0]
                self._log(
                    f"  ✘ Erro de banco em {endpoint}: {msg}"
                )
                erros.append(f"{endpoint}: {msg}")
                erros_detalhados[endpoint] = ("Erro PostgreSQL", exc)

            except requests.HTTPError as exc:
                msg = str(exc)
                self._log(
                    f"  ✘ Erro HTTP em {endpoint}: {msg}"
                )
                erros.append(f"{endpoint}: {msg}")
                erros_detalhados[endpoint] = ("Erro HTTP", exc)

            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                msg = str(exc).splitlines()[0]
                self._log(
                    f"  ✘ Erro em {endpoint}: {msg}"
                )
                erros.append(f"{endpoint}: {msg}")
                erros_detalhados[endpoint] = ("Erro Geral", exc)

            self.progress["value"] = (i / total) * 100
            self.root.update_idletasks()

        try:
            conn.close()
        except Exception:
            pass

        self._log("─" * 60)

        if erros:
            self._log(f"⚠ Concluído com {len(erros)} erro(s):")
            for e in erros:
                self._log(f"  • {e}")
                
            # Gerar um único arquivo consolidado
            caminho_log = registrar_erro(
                "Múltiplos erros durante a importação em lote",
                None,
                **{ep: f"[{tipo}] {str(exc) if exc else ''}" for ep, (tipo, exc) in erros_detalhados.items()}
            )
            
            msg_final = f"Importação finalizada com {len(erros)} erro(s).\n"
            if caminho_log:
                msg_final += f"\nArquivo de erro gerado:\n{caminho_log}\n"
                msg_final += "\nVerifique o log para detalhes."
            else:
                msg_final += "\nVerifique o log para detalhes."
                
            self.root.after(0, lambda: messagebox.showwarning(
                "Concluído com erros",
                msg_final
            ))
            self.root.after(0, lambda: self.btn_executar.config(state="normal"))
        else:
            self._log(cfg["msg_ok"])
            self.root.after(0, self._concluir_fase)


    def _executar_sql_titulos(self):
        """
        Executa a preparação automática dos títulos.
        """
        try:
            conn = psycopg2.connect(**self.conn_params)

            with conn.cursor() as cur:
                cur.execute(SQL_TITULO_PAGAR)

            conn.commit()
            conn.close()
            
            self._log(
                "✔ Preparação dos títulos concluída."
            )

            return True

        except Exception as exc:

            caminho_log = registrar_erro(
                "Erro ao executar SQL_TITULO_PAGAR",
                exc
            )
            self._log(
                f"✘ Erro ao executar SQL: {exc}"
            )

            msg = f"Falha ao executar SQL:\n{exc}"
            if caminho_log:
                msg += f"\n\nArquivo de erro gerado:\n{caminho_log}\n\n"

            self.root.after(
                0,
                lambda m=msg: messagebox.showerror("Erro", m)
            )

            return False
# ==========================================================
# TELA DE TÍTULOS POST (ENVIO PARA API)
# ==========================================================

class TelaTitulosPost:
    """Tela para conversão e envio de Títulos a Pagar para API de destino."""

    # Colunas exibidas: (id_coluna, cabeçalho, largura)
    _COLUNAS = [
        ("fornecedorCodigo",           "Forn. Cód.",         80),
        ("numeroTitulo",               "Nº Título",          90),
        ("descricao",                  "Descrição",         160),
        ("valor",                      "Valor",              80),
        ("parcela",                    "Parcela",            70),
        ("dataMovimento",              "Dt. Movimento",     110),
        ("vencimento",                 "Vencimento",        110),
        ("planoContaGerencialCodigo",  "Plano Cta.",         80),
        ("centroCustoCodigo",          "C. Custo",           80),
        ("observacao",                 "Observação",        140),
        ("isDataValida",               "Dt. Válida",         80),
        ("status",                     "Status",             90),
    ]

    def __init__(self, root, app):
        self.root = root
        self.app = app
        self.conn_params: dict | None = None
        self.titulos: list[dict] = []
        self.frame = tk.Frame(root)
        self.frame.columnconfigure(0, weight=1)
        self.frame.rowconfigure(2, weight=1)
        self._build()

    def _build(self):
        PAD = {"padx": 14, "pady": 8}

        tk.Label(self.frame, text="Importação de Títulos a Pagar para Retaguarda",
                 font=("Segoe UI", 12, "bold")).grid(
                     row=0, column=0, sticky="w", **PAD)

        # Token API de destino
        frame_token = tk.Frame(self.frame)
        frame_token.grid(row=1, column=0, sticky="ew", padx=14, pady=8)
        frame_token.columnconfigure(1, weight=1)

        tk.Label(frame_token, text="Nome da Filial de Destino:", anchor="w").grid(
            row=0, column=0, sticky="w", padx=(0, 10))
        self.entry_token = tk.Entry(frame_token, width=40)
        self.entry_token.grid(row=0, column=1, sticky="ew")
        self.entry_token.config(state="readonly")

        # Frame para a tabela de títulos
        frame_table = tk.Frame(self.frame)
        frame_table.grid(row=2, column=0, sticky="nsew", padx=14, pady=(0, 14))
        frame_table.columnconfigure(0, weight=1)
        frame_table.rowconfigure(0, weight=1)

        # Treeview com scroll vertical e horizontal
        col_ids = [c[0] for c in self._COLUNAS]
        self.tree = ttk.Treeview(frame_table, height=12, columns=col_ids, show="headings")

        for col_id, col_header, col_width in self._COLUNAS:
            self.tree.heading(col_id, text=col_header)
            self.tree.column(col_id, width=col_width, minwidth=50)

        scroll_v = ttk.Scrollbar(frame_table, orient="vertical", command=self.tree.yview)
        scroll_h = ttk.Scrollbar(frame_table, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scroll_v.set, xscrollcommand=scroll_h.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        scroll_v.grid(row=0, column=1, sticky="ns")
        scroll_h.grid(row=1, column=0, sticky="ew")
        frame_table.rowconfigure(0, weight=1)
        frame_table.columnconfigure(0, weight=1)

        # Informações de resumo
        self.lbl_info = tk.Label(self.frame, text="", fg="gray", anchor="w")
        self.lbl_info.grid(row=3, column=0, sticky="ew", padx=14, pady=4)

        # Frame de botões
        frame_btn = tk.Frame(self.frame)
        frame_btn.grid(row=4, column=0, pady=14)
        tk.Button(frame_btn, text="Voltar", command=self._voltar,
                  width=14, relief="flat").pack(side="left", padx=6)
        tk.Button(frame_btn, text="↺ Recarregar", command=self._recarregar,
                  width=14, bg="#B8860B", fg="white", relief="flat").pack(side="left", padx=6)
        self.btn_importar = tk.Button(frame_btn, text="Importar",
                  command=self._iniciar_importacao, width=14,
                  bg="#2E8B57", fg="white", relief="flat")
        self.btn_importar.pack(side="left", padx=6)

    def show(self):
        self.frame.grid(row=0, column=0, sticky="nsew")

    def hide(self):
        self.frame.grid_remove()

    def carregar_titulos(self, conn_params: dict):
        """Carrega os títulos da tabela titulo_pagar_post com todos os campos.

        Pré-preenche o campo de token se o token do destino já estiver salvo no App.
        """
        self.conn_params = conn_params
        self.titulos = []

        # Pré-preenche o nome fantasia da filial de destino
        token_salvo = getattr(self.app, "token_destino", "")
        nome_filial = ""
        if token_salvo:
            try:
                url = f"{BASE_URL}EMPRESAS?chave={token_salvo}"
                resp = requests.get(url, timeout=10)
                resp.raise_for_status()
                dados = resp.json()
                resultados = dados.get("resultados", [])
                if resultados:
                    nome_filial = resultados[0].get("fantasia", "Filial Desconhecida")
                else:
                    nome_filial = "Filial Não Encontrada"
            except Exception as e:
                log.error(f"Erro ao buscar nome da filial: {e}")
                nome_filial = "Erro ao buscar nome da filial"

            self.entry_token.config(state="normal")
            self.entry_token.delete(0, tk.END)
            self.entry_token.insert(0, nome_filial)
            self.entry_token.config(state="readonly")

        # Limpa a tabela
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            conn = psycopg2.connect(**conn_params)
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT
                        fornecedorCodigo,
                        numeroTitulo,
                        descricao,
                        valor,
                        parcela,
                        dataMovimento,
                        vencimento,
                        planoContaGerencialCodigo,
                        centroCustoCodigo,
                        observacao,
                        isDataValida
                    FROM titulo_pagar_post
                    ORDER BY fornecedorCodigo
                    LIMIT 1000
                """)
                rows = cur.fetchall()

            for row in rows:
                # Normaliza chaves para minúsculas — PostgreSQL retorna lowercase
                titulo = {k.lower(): v for k, v in dict(row).items()}
                titulo["status"] = "Pendente"
                self.titulos.append(titulo)
                self.tree.insert("", "end", values=(
                    titulo.get("fornecedorcodigo", ""),
                    titulo.get("numerotitulo", ""),
                    titulo.get("descricao", ""),
                    titulo.get("valor", ""),
                    titulo.get("parcela", ""),
                    titulo.get("datamovimento", ""),
                    titulo.get("vencimento", ""),
                    titulo.get("planocontagerencialcodigo", ""),
                    titulo.get("centrocustocodigo", ""),
                    titulo.get("observacao", ""),
                    titulo.get("isdatavalida", ""),
                    "Pendente",
                ))

            self.lbl_info.config(
                text=f"Total de registros: {len(self.titulos)}"
            )
            conn.close()
        except Exception as exc:
            messagebox.showerror("Erro", f"Erro ao carregar títulos:\n{exc}")
            self.titulos = []

    def _voltar(self):
        """Limpa as informações temporárias e volta para a tela anterior."""
        # Limpa tabela e registros em memória
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.titulos = []
        
        # Limpa mensagens
        self.lbl_info.config(text="")
        
        # Limpa campo visual
        self.entry_token.config(state="normal")
        self.entry_token.delete(0, tk.END)
        self.entry_token.config(state="readonly")
        
        # Restaura botão caso estivesse desabilitado
        self.btn_importar.config(state="normal")

        self.hide()
        self.app.tela_principal.show()

    def _recarregar(self):
        """Recarrega os títulos do banco sem sair da tela."""
        if self.conn_params is None:
            messagebox.showwarning("Sem conexão",
                                   "Parâmetros de conexão não disponíveis.")
            return
        self.lbl_info.config(text="Recarregando...", fg="gray")
        self.root.update_idletasks()
        self.carregar_titulos(self.conn_params)

    def _iniciar_importacao(self):
        """Inicia o processo de importação em thread."""
        token = getattr(self.app, "token_destino", "").strip()
        if not token:
            messagebox.showwarning("Erro",
                                   "Token da API de Destino não encontrado internamente. Volte e tente novamente.")
            return

        if not self.titulos:
            messagebox.showwarning("Sem dados",
                                   "Nenhum título a ser importado.")
            return

        self.btn_importar.config(state="disabled")
        threading.Thread(
            target=self._processar_importacao,
            args=(token,),
            daemon=True,
        ).start()

    def _processar_importacao(self, token: str):
        """Processa a importação dos títulos para a API."""
        sucesso = 0
        erros = 0
        erros_list = []

        try:
            # Busca parâmetros originais de conexão (host/porta/dbname/user/password)
            for i, titulo in enumerate(self.titulos):
                # Converte os campos conforme mapeamento
                payload = self._converter_titulo(titulo)

                try:
                    # Faz POST para a API
                    url = f"{BASE_URL}TITULO_PAGAR?chave={token}"
                    resp = requests.post(url, json=payload, timeout=30)
                    resp.raise_for_status()

                    # Atualiza status
                    titulo["status"] = "✔ Importado"
                    sucesso += 1

                except requests.exceptions.RequestException as exc:
                    registrar_erro(
                        f"Erro ao enviar título {titulo.get('numerotitulo')}",
                        exc
                    )
                    titulo["status"] = f"✘ Erro"
                    msg_erro = str(exc)
                    erros_list.append({
                        "titulo": titulo,
                        "erro": msg_erro,
                    })
                    erros += 1

                # Atualiza a UI
                self.root.after(0, lambda idx=i: self._atualizar_linha(idx))

        except Exception as exc:
            caminho_log = registrar_erro("Erro geral na importação", exc)
            msg = f"Erro geral na importação:\n{exc}"
            if caminho_log:
                msg += f"\n\nArquivo de erro gerado:\n{caminho_log}\n\n"
                
            self.root.after(0, lambda m=msg: messagebox.showerror("Erro", m))
            self.btn_importar.config(state="normal")
            return

        # Resumo final
        total = sucesso + erros
        pct = (sucesso / total * 100) if total > 0 else 0
        resumo = f"Total: {total} | Sucesso: {sucesso} | Erros: {erros} | {pct:.1f}%"

        self.root.after(0, lambda: self._mostrar_resumo(resumo, erros_list))

    def _converter_titulo(self, titulo: dict) -> dict:
        """Converte um título para o formato da API.
        As chaves do dict já estão em minúsculas (normalizadas ao carregar do banco).
        """
        payload = {
            "fornecedorCodigo":  titulo.get("fornecedorcodigo"),
            "numeroTitulo":      str(titulo.get("numerotitulo", "")),
            "descricao":         str(titulo.get("descricao", "")),
            "valorParcela":      float(titulo.get("valor") or 0),
            "numeroParcelas":    int(titulo.get("parcela") or 1),
            "dataMovimento":     self._formatar_data(titulo.get("datamovimento")),
            "vencimento":        self._formatar_data(titulo.get("vencimento")),
            "planoContaCodigo":  titulo.get("planocontagerencialcodigo") or 0,
            "centroCustoCodigo": titulo.get("centrocustocodigo") or 0,
            "observacao":        str(titulo.get("observacao") or ""),
            "isDataValida":      bool(titulo.get("isdatavalida", True)),
        }
        return payload

    def _formatar_data(self, data) -> str:
        """Formata a data para ISO 8601."""
        if data is None:
            return datetime.now().strftime("%Y-%m-%d")
        if isinstance(data, str):
            return data
        if hasattr(data, "strftime"):
            return data.strftime("%Y-%m-%d")
        return datetime.now().strftime("%Y-%m-%d")

    def _atualizar_linha(self, idx: int):
        """Atualiza a visualização de uma linha (chaves já em minúsculas)."""
        if idx < len(self.titulos):
            titulo = self.titulos[idx]
            items = self.tree.get_children()
            if idx < len(items):
                self.tree.item(items[idx], values=(
                    titulo.get("fornecedorcodigo", ""),
                    titulo.get("numerotitulo", ""),
                    titulo.get("descricao", ""),
                    titulo.get("valor", ""),
                    titulo.get("parcela", ""),
                    titulo.get("datamovimento", ""),
                    titulo.get("vencimento", ""),
                    titulo.get("planocontagerencialcodigo", ""),
                    titulo.get("centrocustocodigo", ""),
                    titulo.get("observacao", ""),
                    titulo.get("isdatavalida", ""),
                    titulo.get("status", ""),
                ))

    def _mostrar_resumo(self, resumo: str, erros_list: list):
        """Exibe o resumo da importação, gera Excel se houver erros e pergunta se deseja voltar."""
        cor = "red" if erros_list else "green"
        self.lbl_info.config(text=resumo, fg=cor)
        self.btn_importar.config(state="normal")

        # ── Gera Excel de erros automaticamente ───────────────────────
        if erros_list:
            if EXCEL_DISPONIVEL:
                caminho = self._gerar_excel_erros(erros_list)
                if caminho:
                    messagebox.showwarning(
                        "Registros rejeitados",
                        f"{len(erros_list)} registro(s) rejeitado(s) pela API.\n\n"
                        f"Log de erros salvo em:\n{caminho}",
                    )
            else:
                messagebox.showwarning(
                    "Aviso",
                    f"Foram encontrados {len(erros_list)} erro(s),\n"
                    "mas openpyxl não está instalado para gerar Excel.",
                )

        # ── Mensagem de conclusão + pergunta ao usuário ────────────────
        total_str = resumo  # já contém "Total: X | Sucesso: Y | Erros: Z | 99.9%"
        if erros_list:
            msg_conclusao = (
                f"✔ Importação concluída com erros!\n\n"
                f"{total_str}\n\n"
                "Deseja voltar à tela de conexão?"
            )
        else:
            msg_conclusao = (
                f"✔ Importação concluída com sucesso!\n\n"
                f"{total_str}\n\n"
                "Deseja voltar à tela de conexão?"
            )

        voltar = messagebox.askyesno("Importação finalizada", msg_conclusao)
        if voltar:
            self.hide()
            self.app.show_conexao()

    def _gerar_excel_erros(self, erros_list: list) -> str | None:
        """
        Gera automaticamente um arquivo Excel com TODOS os campos originais
        do registro + coluna ERRO_IMPORTACAO ao final.

        Salva em: C:\\Quality\\LOG\\erros_implantacao\\
        Retorna o caminho completo do arquivo gerado, ou None em caso de falha.
        """
        PASTA_LOG = r"C:\Quality\LOG\erros_implantacao"

        try:
            os.makedirs(PASTA_LOG, exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = "Erros Importação"

            # Colunas dos dados (mesma ordem da tela e do mapeamento)
            COLUNAS_DADOS = [
                "fornecedorcodigo",
                "numerotitulo",
                "descricao",
                "valor",
                "parcela",
                "datamovimento",
                "vencimento",
                "planocontagerencialcodigo",
                "centrocustocodigo",
                "observacao",
                "isdatavalida",
            ]
            # Cabeçalhos legíveis para o Excel
            CABECALHOS = [
                "fornecedorCodigo",
                "numeroTitulo",
                "descricao",
                "valor",
                "parcela",
                "dataMovimento",
                "vencimento",
                "planoContaGerencialCodigo",
                "centroCustoCodigo",
                "observacao",
                "isDataValida",
                "ERRO_IMPORTACAO",          # coluna extra ao final
            ]

            # ── Estilos ────────────────────────────────────────────────
            fill_normal = PatternFill(
                start_color="1F3864", end_color="1F3864", fill_type="solid"
            )
            fill_erro = PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            )
            font_header = Font(bold=True, color="FFFFFF", size=10)
            font_erro_col = Font(bold=True, color="FFFFFF", size=10)
            from openpyxl.styles import Alignment, Border, Side
            thin = Side(border_style="thin", color="AAAAAA")
            borda = Border(left=thin, right=thin, top=thin, bottom=thin)
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

            # ── Cabeçalho ──────────────────────────────────────────────
            ws.append(CABECALHOS)
            ws.row_dimensions[1].height = 28

            for col_idx, cell in enumerate(ws[1], start=1):
                is_erro_col = (col_idx == len(CABECALHOS))
                cell.fill   = fill_erro if is_erro_col else fill_normal
                cell.font   = font_erro_col if is_erro_col else font_header
                cell.border = borda
                cell.alignment = align_center

            # ── Dados ──────────────────────────────────────────────────
            fill_row_par   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            fill_row_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            font_erro_msg  = Font(color="C00000", bold=True)

            for linha_idx, erro_info in enumerate(erros_list, start=2):
                titulo    = erro_info["titulo"]
                erro_msg  = erro_info["erro"]

                linha = [titulo.get(col, "") for col in COLUNAS_DADOS]
                linha.append(erro_msg)
                ws.append(linha)

                fill_row = fill_row_par if linha_idx % 2 == 0 else fill_row_impar
                for col_idx, cell in enumerate(ws[linha_idx], start=1):
                    cell.border    = borda
                    cell.alignment = align_left
                    if col_idx == len(CABECALHOS):
                        cell.font = font_erro_msg
                    else:
                        cell.fill = fill_row

            # ── Largura automática das colunas ─────────────────────────
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value or "")) for cell in col),
                    default=10,
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            # ── Congela cabeçalho ──────────────────────────────────────
            ws.freeze_panes = "A2"

            # ── Salva ──────────────────────────────────────────────────
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"erros_importacao_{timestamp}.xlsx"
            caminho_completo = os.path.join(PASTA_LOG, nome_arquivo)
            wb.save(caminho_completo)

            log.info(f"Excel de erros salvo em: {caminho_completo}")
            return caminho_completo

        except Exception as exc:
            log.exception("Falha ao gerar Excel de erros")
            messagebox.showerror(
                "Erro ao gerar Excel",
                f"Não foi possível salvar o log de erros:\n{exc}",
            )
            return None


def buscar_dados(endpoint: str, chave: str, log_fn=None) -> list[dict]:
    """Busca todos os registros via paginação por ultimoCodigo."""
    _log = log_fn or log.info
    registros: list[dict] = []
    ultimo_codigo = 0

    while True:
        url = (
            f"{BASE_URL}{endpoint}"
            f"?CHAVE={chave}&ultimoCodigo={ultimo_codigo}"
        )
        log.info(f"GET {url}")

        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        data = resp.json()

        pagina      = data.get("resultados", [])
        novo_ultimo = data.get("ultimoCodigo", ultimo_codigo)

        registros.extend(pagina)

        if not pagina or novo_ultimo <= ultimo_codigo:
            break

        ultimo_codigo = novo_ultimo
        _log(f"  Página: +{len(pagina)} registros (acumulado: {len(registros)})")

    return registros


# ==========================================================
# BANCO DE DADOS
# ==========================================================

def inserir_no_banco(
    conn,
    tabela: str,
    registros: list[dict],
    log_fn=None,
) -> int:
    """
    Insere registros na tabela PostgreSQL.

    - Consulta as colunas REAIS do banco via information_schema (case-insensitive).
    - Filtra apenas as colunas definidas em COLUNAS_ESPERADAS que existem no banco.
    - Faz rollback isolado em caso de erro (não contamina os próximos endpoints).
    - Usa ON CONFLICT DO NOTHING para idempotência.
    """
    _log = log_fn or log.info

    if not registros:
        return 0

    colunas_api = COLUNAS_ESPERADAS.get(tabela)
    if not colunas_api:
        raise ValueError(
            f"Tabela '{tabela}' não possui definição em COLUNAS_ESPERADAS."
        )

    # Mapeamento: nome_api → nome_real_no_banco
    mapeamento = mapear_colunas(conn, tabela, colunas_api, log_fn=_log)

    if not mapeamento:
        raise ValueError(
            f"Nenhuma coluna correspondente encontrada no banco para '{tabela}'."
        )

    nomes_api   = [m[0] for m in mapeamento]
    nomes_banco = [m[1] for m in mapeamento]

    col_str      = ", ".join(f'"{c}"' for c in nomes_banco)
    placeholders = ", ".join(["%s"] * len(nomes_banco))
    sql = (
        f'INSERT INTO "{tabela}" ({col_str}) '
        f'VALUES ({placeholders}) '
        f'ON CONFLICT DO NOTHING'
    )

    linhas = [
        tuple(
            _preparar_valor(r.get(col_api), col_api)
            for col_api in nomes_api
        )
        for r in registros
    ]

    try:
        with conn.cursor() as cur:
            psycopg2.extras.execute_batch(cur, sql, linhas, page_size=500)
            # execute_batch não garante rowcount confiável — usa len como fallback
            inseridos = cur.rowcount if cur.rowcount >= 0 else len(linhas)
        conn.commit()
        return inseridos

    except Exception:
        conn.rollback()   # isola o erro — próximo endpoint começa limpo
        raise


def _preparar_valor(valor, coluna: str):
    """
    Prepara o valor para inserção no banco.
    - Colunas JSONB: usa psycopg2.extras.Json
    - Outros dict/list: serializa como string JSON
    - Demais: passa direto
    """
    if valor is None:
        return None

    col_lower = coluna.lower()

    if col_lower in {c.lower() for c in COLUNAS_JSONB}:
        if isinstance(valor, (dict, list)):
            return psycopg2.extras.Json(valor)
        try:
            return psycopg2.extras.Json(json.loads(valor))
        except (TypeError, ValueError):
            return psycopg2.extras.Json(valor)

    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False)

    return valor

def registrar_erro(mensagem: str, excecao: Exception = None, **kwargs):
    """
    Salva erro no arquivo de log e gera um arquivo txt detalhado individual.
    """
    if excecao:
        log.exception(f"{mensagem}: {excecao}")
    else:
        log.error(mensagem)
        
    try:
        agora = datetime.now()
        timestamp = agora.strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"erro_titulopagar_{timestamp}.txt"
        caminho_arquivo = os.path.join(PASTA_LOG_ERROS, nome_arquivo)
        
        # Tenta descobrir o módulo/função que chamou
        caller_frame = inspect.stack()[1]
        modulo = os.path.basename(caller_frame.filename)
        operacao = caller_frame.function
        
        versao_app = "1.0.0"
        
        linhas_erro = []
        linhas_erro.append("=" * 60)
        linhas_erro.append(" RELATÓRIO DE ERRO - TÍTULOS A PAGAR")
        linhas_erro.append("=" * 60)
        linhas_erro.append(f"Data e Hora      : {agora.strftime('%d/%m/%Y %H:%M:%S')}")
        linhas_erro.append(f"Versão Aplicação : {versao_app}")
        linhas_erro.append(f"Módulo/Tela      : {modulo}")
        linhas_erro.append(f"Operação         : {operacao}")
        linhas_erro.append(f"Contexto/Msg     : {mensagem}")
        
        if kwargs:
            linhas_erro.append("-" * 60)
            linhas_erro.append(" DADOS ADICIONAIS:")
            for k, v in kwargs.items():
                linhas_erro.append(f"  {k}: {v}")
                
        linhas_erro.append("-" * 60)
        if excecao:
            linhas_erro.append(f"Exceção: {type(excecao).__name__} - {str(excecao)}")
            linhas_erro.append("Stack Trace:")
            linhas_erro.append("".join(traceback.format_exception(type(excecao), excecao, excecao.__traceback__)))
        else:
            linhas_erro.append("Stack Trace da Chamada (Sem exceção explícita):")
            linhas_erro.append("".join(traceback.format_stack()[:-1]))
            
        linhas_erro.append("=" * 60)
        
        with open(caminho_arquivo, "w", encoding="utf-8") as f:
            f.write("\n".join(linhas_erro))
            
        return caminho_arquivo
            
    except Exception as e:
        log.error(f"Falha ao gerar arquivo de erro detalhado: {e}")
        return None
# ==========================================================
# UTILITÁRIOS
# ==========================================================

def _mensagem_amigavel(exc: Exception) -> str:
    msg = str(exc).lower()
    if "password" in msg or "authentication" in msg:
        return "Senha incorreta ou usuário sem permissão."
    if "connection refused" in msg:
        return "Conexão recusada — verifique host, porta e se o servidor está ativo."
    if "does not exist" in msg:
        return "Banco de dados não encontrado."
    if "timeout" in msg:
        return "Tempo de conexão esgotado — verifique o host e a rede."
    if "name or service not known" in msg or "nodename" in msg:
        return "Host não encontrado — verifique o endereço IP/hostname."
    return str(exc)


# ==========================================================
# MAIN
# ==========================================================

def main() -> None:
    app = App()
    app.run()


if __name__ == "__main__":
    main()